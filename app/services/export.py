"""PDF and Excel export for financial reports."""
import io
import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import arabic_reshaper
from bidi.algorithm import get_display
from app.services.reports import (
    balance_sheet, income_statement, cash_flow,
    income_summary, expenses_summary, income_statement_compared,
    aging_report, ap_aging_report, vat_report,
    payroll_summary_report, fixed_assets_report,
)

NAVY = colors.HexColor("#0A2540")
BLUE = colors.HexColor("#2563EB")
GRAY = colors.HexColor("#64748B")

# ─── Arabic-capable font registration ──────────────────────────────────
_FONT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "static", "fonts")
_FONT_REGULAR = "Amiri"
_FONT_BOLD = "Amiri-Bold"


def _register_fonts():
    if _FONT_REGULAR in pdfmetrics.getRegisteredFontNames():
        return
    pdfmetrics.registerFont(TTFont(_FONT_REGULAR, os.path.join(_FONT_DIR, "Amiri-Regular.ttf")))
    pdfmetrics.registerFont(TTFont(_FONT_BOLD, os.path.join(_FONT_DIR, "Amiri-Bold.ttf")))


_register_fonts()


def ar(text):
    """Shape Arabic text for correct rendering in reportlab PDFs.

    Safe to call on any string: reshapes Arabic ligatures + applies bidi
    for correct visual order. Pure-Latin strings pass through unchanged.
    """
    if text is None:
        return ""
    s = str(text)
    if not s:
        return s
    reshaped = arabic_reshaper.reshape(s)
    return get_display(reshaped)


def _excel_styled_header(ws, title, company_name, period):
    ws["A1"] = company_name
    ws["A1"].font = Font(size=16, bold=True, color="0A2540")
    ws["A2"] = title
    ws["A2"].font = Font(size=14, bold=True, color="2563EB")
    ws["A3"] = period
    ws["A3"].font = Font(size=10, color="64748B")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18


def export_balance_sheet_excel(company, as_of):
    data = balance_sheet(company.id, as_of=as_of)
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"
    _excel_styled_header(ws, "Balance Sheet — الميزانية العمومية", company.name, f"كما في {as_of}")

    row = 5
    ws.cell(row=row, column=1, value="الأصول").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="0A2540")
    row += 1
    for item in data["assets"]:
        ws.cell(row=row, column=1, value=f"  {item['code']} — {item['name']}")
        ws.cell(row=row, column=2, value=item["balance"]).number_format = "#,##0.00"
        row += 1
    ws.cell(row=row, column=1, value="إجمالي الأصول").font = Font(bold=True)
    ws.cell(row=row, column=2, value=data["totals"]["assets"]).font = Font(bold=True)
    ws.cell(row=row, column=2).number_format = "#,##0.00"
    row += 2

    ws.cell(row=row, column=1, value="الالتزامات").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="0A2540")
    row += 1
    for item in data["liabilities"]:
        ws.cell(row=row, column=1, value=f"  {item['code']} — {item['name']}")
        ws.cell(row=row, column=2, value=item["balance"]).number_format = "#,##0.00"
        row += 1
    ws.cell(row=row, column=1, value="إجمالي الالتزامات").font = Font(bold=True)
    ws.cell(row=row, column=2, value=data["totals"]["liabilities"]).font = Font(bold=True)
    row += 2

    ws.cell(row=row, column=1, value="حقوق الملكية").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="0A2540")
    row += 1
    for item in data["equity"]:
        ws.cell(row=row, column=1, value=f"  {item['code']} — {item['name']}")
        ws.cell(row=row, column=2, value=item["balance"]).number_format = "#,##0.00"
        row += 1
    ws.cell(row=row, column=1, value="إجمالي حقوق الملكية").font = Font(bold=True)
    ws.cell(row=row, column=2, value=data["totals"]["equity"]).font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_income_statement_excel(company, start, end):
    data = income_statement(company.id, start_date=start, end_date=end)
    wb = Workbook()
    ws = wb.active
    ws.title = "Income Statement"
    _excel_styled_header(ws, "Income Statement — قائمة الدخل", company.name, f"من {start} إلى {end}")

    row = 5
    ws.cell(row=row, column=1, value="الإيرادات").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="0A2540")
    row += 1
    for item in data["revenue"]:
        ws.cell(row=row, column=1, value=f"  {item['code']} — {item['name']}")
        ws.cell(row=row, column=2, value=item["balance"]).number_format = "#,##0.00"
        row += 1
    ws.cell(row=row, column=1, value="إجمالي الإيرادات").font = Font(bold=True)
    ws.cell(row=row, column=2, value=data["total_revenue"]).font = Font(bold=True)
    row += 2

    ws.cell(row=row, column=1, value="المصروفات").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="0A2540")
    row += 1
    for item in data["expenses"]:
        ws.cell(row=row, column=1, value=f"  {item['code']} — {item['name']}")
        ws.cell(row=row, column=2, value=item["balance"]).number_format = "#,##0.00"
        row += 1
    ws.cell(row=row, column=1, value="إجمالي المصروفات").font = Font(bold=True)
    ws.cell(row=row, column=2, value=data["total_expense"]).font = Font(bold=True)
    row += 2

    color = "10B981" if data["net_income"] >= 0 else "EF4444"
    ws.cell(row=row, column=1, value="صافي الربح / الخسارة").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=color)
    ws.cell(row=row, column=2, value=data["net_income"]).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=color)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _pdf_header(p, company, title, period):
    p.setFillColor(NAVY)
    p.rect(0, 27.7 * cm, 21 * cm, 2 * cm, fill=1, stroke=0)
    p.setFillColor(colors.white)
    p.setFont(_FONT_BOLD, 18)
    p.drawString(1.5 * cm, 28.5 * cm, ar(company.name))
    p.setFont(_FONT_REGULAR, 10)
    p.drawString(1.5 * cm, 28 * cm, ar("Marsoud — Financial Report"))

    p.setFillColor(NAVY)
    p.setFont(_FONT_BOLD, 16)
    p.drawString(1.5 * cm, 26.5 * cm, ar(title))
    p.setFillColor(GRAY)
    p.setFont(_FONT_REGULAR, 10)
    p.drawString(1.5 * cm, 26 * cm, ar(period))


def _pdf_section(p, y, label):
    p.setFillColor(NAVY)
    p.rect(1 * cm, y - 0.4 * cm, 19 * cm, 0.7 * cm, fill=1, stroke=0)
    p.setFillColor(colors.white)
    p.setFont(_FONT_BOLD, 11)
    p.drawString(1.3 * cm, y - 0.2 * cm, ar(label))
    return y - 1 * cm


def export_balance_sheet_pdf(company, as_of):
    data = balance_sheet(company.id, as_of=as_of)
    buf = io.BytesIO()
    p = canvas.Canvas(buf, pagesize=A4)
    _pdf_header(p, company, "Balance Sheet", f"As of {as_of}")

    y = 24.5 * cm
    for section, items, total_key in [
        ("ASSETS", data["assets"], "assets"),
        ("LIABILITIES", data["liabilities"], "liabilities"),
        ("EQUITY", data["equity"], "equity"),
    ]:
        y = _pdf_section(p, y, section)
        p.setFillColor(colors.black)
        p.setFont(_FONT_REGULAR, 10)
        for item in items:
            if y < 3 * cm:
                p.showPage()
                _pdf_header(p, company, "Balance Sheet (cont.)", f"As of {as_of}")
                y = 24.5 * cm
            p.drawString(1.5 * cm, y, ar(f"{item['code']}  {item['name']}"))
            p.drawRightString(19.5 * cm, y, f"{item['balance']:,.2f}")
            y -= 0.5 * cm
        p.setFont(_FONT_BOLD, 10)
        p.setFillColor(BLUE)
        p.drawString(1.5 * cm, y, ar(f"Total {section}"))
        p.drawRightString(19.5 * cm, y, f"{data['totals'][total_key]:,.2f}")
        p.setFillColor(colors.black)
        y -= 1 * cm

    p.showPage()
    p.save()
    buf.seek(0)
    return buf


def export_income_statement_pdf(company, start, end):
    data = income_statement(company.id, start_date=start, end_date=end)
    buf = io.BytesIO()
    p = canvas.Canvas(buf, pagesize=A4)
    _pdf_header(p, company, "Income Statement", f"{start} → {end}")

    y = 24.5 * cm
    for section, items, total_key in [
        ("REVENUE", data["revenue"], "total_revenue"),
        ("EXPENSES", data["expenses"], "total_expense"),
    ]:
        y = _pdf_section(p, y, section)
        p.setFillColor(colors.black)
        p.setFont(_FONT_REGULAR, 10)
        for item in items:
            p.drawString(1.5 * cm, y, ar(f"{item['code']}  {item['name']}"))
            p.drawRightString(19.5 * cm, y, f"{item['balance']:,.2f}")
            y -= 0.5 * cm
        p.setFont(_FONT_BOLD, 10)
        p.setFillColor(BLUE)
        p.drawString(1.5 * cm, y, ar(f"Total {section}"))
        p.drawRightString(19.5 * cm, y, f"{data[total_key]:,.2f}")
        p.setFillColor(colors.black)
        y -= 1 * cm

    y -= 0.3 * cm
    profit = data["net_income"]
    color = colors.HexColor("#10B981") if profit >= 0 else colors.HexColor("#EF4444")
    p.setFillColor(color)
    p.rect(1 * cm, y - 0.5 * cm, 19 * cm, 0.9 * cm, fill=1, stroke=0)
    p.setFillColor(colors.white)
    p.setFont(_FONT_BOLD, 13)
    p.drawString(1.5 * cm, y - 0.2 * cm, "NET PROFIT / (LOSS)")
    p.drawRightString(19.5 * cm, y - 0.2 * cm, f"{profit:,.2f}")

    p.showPage()
    p.save()
    buf.seek(0)
    return buf


def export_invoice_pdf(invoice):
    """Generate a customer-facing invoice PDF."""
    buf = io.BytesIO()
    p = canvas.Canvas(buf, pagesize=A4)
    _pdf_header(p, invoice.company, f"Invoice {invoice.number}", f"Date: {invoice.issue_date}  ·  Due: {invoice.due_date}")

    y = 24.5 * cm
    p.setFillColor(colors.HexColor("#475569"))
    p.setFont(_FONT_BOLD, 11)
    p.drawString(1.5 * cm, y, "BILL TO:")
    y -= 0.5 * cm
    p.setFillColor(colors.black)
    p.setFont(_FONT_REGULAR, 11)
    p.drawString(1.5 * cm, y, ar(invoice.customer.name))
    if invoice.customer.email:
        y -= 0.5 * cm
        p.setFont(_FONT_REGULAR, 9)
        p.drawString(1.5 * cm, y, invoice.customer.email)

    y -= 1.2 * cm
    # Items header
    p.setFillColor(NAVY)
    p.rect(1 * cm, y - 0.4 * cm, 19 * cm, 0.7 * cm, fill=1, stroke=0)
    p.setFillColor(colors.white)
    p.setFont(_FONT_BOLD, 10)
    p.drawString(1.3 * cm, y - 0.15 * cm, "DESCRIPTION")
    p.drawString(12 * cm, y - 0.15 * cm, "QTY")
    p.drawString(14 * cm, y - 0.15 * cm, "PRICE")
    p.drawRightString(19.5 * cm, y - 0.15 * cm, "TOTAL")
    y -= 1 * cm

    p.setFillColor(colors.black)
    p.setFont(_FONT_REGULAR, 10)
    for item in invoice.items:
        p.drawString(1.3 * cm, y, ar(item.description[:60]))
        p.drawString(12 * cm, y, f"{float(item.quantity):.2f}")
        p.drawString(14 * cm, y, f"{float(item.unit_price):,.2f}")
        p.drawRightString(19.5 * cm, y, f"{item.total:,.2f}")
        y -= 0.5 * cm

    y -= 0.5 * cm
    p.setFont(_FONT_REGULAR, 10)
    p.drawRightString(17 * cm, y, "Subtotal:")
    p.drawRightString(19.5 * cm, y, f"{float(invoice.subtotal):,.2f}")
    y -= 0.5 * cm
    p.drawRightString(17 * cm, y, f"VAT ({float(invoice.tax_rate):.0f}%):")
    p.drawRightString(19.5 * cm, y, f"{float(invoice.tax_amount):,.2f}")
    y -= 0.7 * cm
    p.setFillColor(NAVY)
    p.rect(13 * cm, y - 0.4 * cm, 7 * cm, 0.8 * cm, fill=1, stroke=0)
    p.setFillColor(colors.white)
    p.setFont(_FONT_BOLD, 12)
    p.drawString(13.3 * cm, y - 0.15 * cm, "TOTAL:")
    p.drawRightString(19.5 * cm, y - 0.15 * cm, f"{float(invoice.total):,.2f} {invoice.currency}")

    p.showPage()
    p.save()
    buf.seek(0)
    return buf


def export_payslip_pdf(employee, line, run):
    """One-page payslip for an employee."""
    buf = io.BytesIO()
    p = canvas.Canvas(buf, pagesize=A4)
    _pdf_header(p, employee.company, f"Payslip — {run.period_month}/{run.period_year}",
                f"{employee.name}  ·  {employee.job_title or ''}")

    y = 24 * cm
    p.setFillColor(colors.HexColor("#475569"))
    p.setFont(_FONT_BOLD, 10)
    p.drawString(1.5 * cm, y, ar(f"Employee #:  {employee.employee_number or '—'}"))
    y -= 0.5 * cm
    p.drawString(1.5 * cm, y, ar(f"Working days: {line.working_days}/30"))
    y -= 1.5 * cm

    rows = [
        ("Basic (prorated)", float(line.basic), False),
        ("Allowances", float(line.allowances), False),
        ("Overtime", float(line.overtime), False),
        ("Bonus", float(line.bonus), False),
        ("Fixed deductions", -float(line.deductions), True),
        ("Absence", -float(line.absence_deduction), True),
        ("Late", -float(line.late_deduction), True),
        ("Advance", -float(line.advance_deduction), True),
    ]
    p.setFont(_FONT_REGULAR, 10)
    for label, value, neg in rows:
        if abs(value) < 0.01:
            continue
        p.setFillColor(colors.HexColor("#B91C1C") if neg else colors.black)
        p.drawString(1.5 * cm, y, ar(label))
        p.drawRightString(19.5 * cm, y, f"{value:+,.2f}")
        y -= 0.55 * cm

    y -= 0.5 * cm
    p.setFillColor(NAVY)
    p.rect(1 * cm, y - 0.5 * cm, 19 * cm, 1 * cm, fill=1, stroke=0)
    p.setFillColor(colors.white)
    p.setFont(_FONT_BOLD, 13)
    p.drawString(1.5 * cm, y - 0.15 * cm, "NET PAY")
    p.drawRightString(19.5 * cm, y - 0.15 * cm, f"{float(line.net):,.2f}")

    p.showPage()
    p.save()
    buf.seek(0)
    return buf


def export_journal_entry_pdf(entry):
    """Single journal entry PDF with all its lines."""
    buf = io.BytesIO()
    p = canvas.Canvas(buf, pagesize=A4)
    title = f"Journal Entry {entry.number or '#' + str(entry.id)}"
    _pdf_header(p, entry.company, title, f"Date: {entry.date}  ·  {entry.currency}")

    y = 24 * cm
    p.setFillColor(colors.HexColor("#475569"))
    p.setFont(_FONT_REGULAR, 10)
    p.drawString(1.5 * cm, y, ar(f"Description: {entry.description}"))
    y -= 0.5 * cm
    if entry.reference:
        p.drawString(1.5 * cm, y, ar(f"Reference: {entry.reference}"))
        y -= 0.5 * cm
    p.drawString(1.5 * cm, y, ar(f"Status: {'Active' if entry.is_active else 'PAUSED'}"))
    y -= 1 * cm

    p.setFillColor(NAVY)
    p.rect(1 * cm, y - 0.4 * cm, 19 * cm, 0.7 * cm, fill=1, stroke=0)
    p.setFillColor(colors.white)
    p.setFont(_FONT_BOLD, 10)
    p.drawString(1.3 * cm, y - 0.2 * cm, "ACCOUNT")
    p.drawString(11 * cm, y - 0.2 * cm, "MEMO")
    p.drawString(15.5 * cm, y - 0.2 * cm, "DEBIT")
    p.drawRightString(19.5 * cm, y - 0.2 * cm, "CREDIT")
    y -= 0.9 * cm

    p.setFillColor(colors.black)
    p.setFont(_FONT_REGULAR, 9)
    for line in entry.lines:
        if y < 3 * cm:
            p.showPage()
            y = 27 * cm
        acc_label = f"{line.account.code} {line.account.name[:30]}"
        p.drawString(1.3 * cm, y, ar(acc_label))
        p.drawString(11 * cm, y, ar((line.memo or "")[:25]))
        p.drawString(15.5 * cm, y, f"{float(line.debit):,.2f}")
        p.drawRightString(19.5 * cm, y, f"{float(line.credit):,.2f}")
        y -= 0.5 * cm

    y -= 0.3 * cm
    p.setFillColor(NAVY)
    p.rect(1 * cm, y - 0.5 * cm, 19 * cm, 0.8 * cm, fill=1, stroke=0)
    p.setFillColor(colors.white)
    p.setFont(_FONT_BOLD, 11)
    p.drawString(1.3 * cm, y - 0.15 * cm, "TOTAL")
    p.drawString(15.5 * cm, y - 0.15 * cm, f"{entry.total_debit:,.2f}")
    p.drawRightString(19.5 * cm, y - 0.15 * cm, f"{entry.total_credit:,.2f}")

    p.showPage()
    p.save()
    buf.seek(0)
    return buf


def export_journal_entry_excel(entry):
    """Single journal entry Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = entry.number or f"JE-{entry.id}"
    _excel_styled_header(ws, f"Journal {entry.number or entry.id}",
                         entry.company.name, f"Date: {entry.date}")

    row = 5
    ws.cell(row=row, column=1, value="Description:").font = Font(bold=True)
    ws.cell(row=row, column=2, value=entry.description)
    row += 1
    if entry.reference:
        ws.cell(row=row, column=1, value="Reference:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=entry.reference)
        row += 1
    row += 1

    for col, h in enumerate(["Account Code", "Account Name", "Memo", "Debit", "Credit"], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0A2540")
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    row += 1
    for line in entry.lines:
        ws.cell(row=row, column=1, value=line.account.code)
        ws.cell(row=row, column=2, value=line.account.name_ar or line.account.name)
        ws.cell(row=row, column=3, value=line.memo or "")
        ws.cell(row=row, column=4, value=float(line.debit)).number_format = "#,##0.00"
        ws.cell(row=row, column=5, value=float(line.credit)).number_format = "#,##0.00"
        row += 1

    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True, color="FFFFFF")
    for col in range(1, 6):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="2563EB")
    ws.cell(row=row, column=4, value=entry.total_debit).number_format = "#,##0.00"
    ws.cell(row=row, column=5, value=entry.total_credit).number_format = "#,##0.00"
    ws.cell(row=row, column=4).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=row, column=5).font = Font(bold=True, color="FFFFFF")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_journals_list_pdf(company, entries, period_label=""):
    """Filtered list of journals to PDF — one row per entry."""
    buf = io.BytesIO()
    p = canvas.Canvas(buf, pagesize=A4)
    _pdf_header(p, company, "Journal Entries List", period_label or "All filtered entries")

    y = 24.5 * cm
    p.setFillColor(NAVY)
    p.rect(1 * cm, y - 0.4 * cm, 19 * cm, 0.7 * cm, fill=1, stroke=0)
    p.setFillColor(colors.white)
    p.setFont(_FONT_BOLD, 9)
    p.drawString(1.3 * cm, y - 0.2 * cm, "#")
    p.drawString(3 * cm, y - 0.2 * cm, "DATE")
    p.drawString(5.5 * cm, y - 0.2 * cm, "DESCRIPTION")
    p.drawString(12 * cm, y - 0.2 * cm, "REF")
    p.drawString(14 * cm, y - 0.2 * cm, "STATUS")
    p.drawString(15.5 * cm, y - 0.2 * cm, "DEBIT")
    p.drawRightString(19.5 * cm, y - 0.2 * cm, "CREDIT")
    y -= 0.8 * cm

    p.setFillColor(colors.black)
    p.setFont(_FONT_REGULAR, 8)
    total_d = total_c = 0.0
    for e in entries:
        if y < 3 * cm:
            p.showPage()
            y = 27 * cm
        p.drawString(1.3 * cm, y, ar((e.number or f"#{e.id}")[:10]))
        p.drawString(3 * cm, y, str(e.date))
        p.drawString(5.5 * cm, y, ar((e.description or "")[:42]))
        p.drawString(12 * cm, y, ar((e.reference or "")[:10]))
        p.drawString(14 * cm, y, ar("Active" if e.is_active else "PAUSED"))
        p.drawString(15.5 * cm, y, f"{e.total_debit:,.2f}")
        p.drawRightString(19.5 * cm, y, f"{e.total_credit:,.2f}")
        total_d += e.total_debit
        total_c += e.total_credit
        y -= 0.45 * cm

    y -= 0.3 * cm
    p.setFillColor(NAVY)
    p.rect(1 * cm, y - 0.5 * cm, 19 * cm, 0.8 * cm, fill=1, stroke=0)
    p.setFillColor(colors.white)
    p.setFont(_FONT_BOLD, 10)
    p.drawString(1.3 * cm, y - 0.15 * cm, ar(f"TOTAL ({len(entries)} entries)"))
    p.drawString(15.5 * cm, y - 0.15 * cm, f"{total_d:,.2f}")
    p.drawRightString(19.5 * cm, y - 0.15 * cm, f"{total_c:,.2f}")

    p.showPage()
    p.save()
    buf.seek(0)
    return buf


def export_payroll_run_pdf(run):
    """Full monthly payroll run PDF — one row per employee."""
    buf = io.BytesIO()
    p = canvas.Canvas(buf, pagesize=A4)
    _pdf_header(p, run.company, f"Payroll — {run.period_month}/{run.period_year}",
                f"{run.number or ''}")

    y = 24.5 * cm
    # Column header
    p.setFillColor(NAVY)
    p.rect(1 * cm, y - 0.4 * cm, 19 * cm, 0.7 * cm, fill=1, stroke=0)
    p.setFillColor(colors.white)
    p.setFont(_FONT_BOLD, 9)
    p.drawString(1.3 * cm, y - 0.2 * cm, "EMPLOYEE")
    p.drawString(7.5 * cm, y - 0.2 * cm, "DAYS")
    p.drawString(9 * cm, y - 0.2 * cm, "BASIC")
    p.drawString(11.3 * cm, y - 0.2 * cm, "ALLOW")
    p.drawString(13.3 * cm, y - 0.2 * cm, "BONUS")
    p.drawString(15.3 * cm, y - 0.2 * cm, "DEDUCT")
    p.drawRightString(19.5 * cm, y - 0.2 * cm, "NET")
    y -= 0.9 * cm

    p.setFillColor(colors.black)
    p.setFont(_FONT_REGULAR, 9)
    for line in run.lines:
        if y < 3 * cm:
            p.showPage()
            y = 27 * cm
        p.drawString(1.3 * cm, y, ar(line.employee.name[:30]))
        p.drawString(7.5 * cm, y, f"{line.working_days}/30")
        p.drawString(9 * cm, y, f"{float(line.basic):,.2f}")
        p.drawString(11.3 * cm, y, f"{float(line.allowances):,.2f}")
        bonus_total = float(line.overtime or 0) + float(line.bonus or 0)
        p.drawString(13.3 * cm, y, f"{bonus_total:,.2f}")
        deduct_total = float(line.deductions or 0) + float(line.absence_deduction or 0) + float(line.late_deduction or 0) + float(line.advance_deduction or 0)
        p.drawString(15.3 * cm, y, f"{deduct_total:,.2f}")
        p.drawRightString(19.5 * cm, y, f"{float(line.net):,.2f}")
        y -= 0.5 * cm

    y -= 0.3 * cm
    p.setFillColor(NAVY)
    p.rect(1 * cm, y - 0.5 * cm, 19 * cm, 0.9 * cm, fill=1, stroke=0)
    p.setFillColor(colors.white)
    p.setFont(_FONT_BOLD, 12)
    p.drawString(1.3 * cm, y - 0.15 * cm, "TOTAL NET")
    p.drawRightString(19.5 * cm, y - 0.15 * cm, f"{float(run.total_net):,.2f} {run.company.base_currency}")

    p.showPage()
    p.save()
    buf.seek(0)
    return buf


def export_payroll_run_excel(run):
    """Full monthly payroll run as Excel — auditable detail per employee."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"{run.period_year}-{run.period_month:02d}"
    _excel_styled_header(ws, f"Payroll {run.period_month}/{run.period_year}",
                         run.company.name, run.number or "")

    headers = ["Employee #", "Name", "Days", "Basic (prorated)", "Allowances",
               "Overtime", "Bonus", "Fixed Deductions", "Absence", "Late", "Advance", "Net"]
    row = 5
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0A2540")
        c.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    for col_letter in "CDEFGHIJKL":
        ws.column_dimensions[col_letter].width = 14

    row += 1
    for line in run.lines:
        emp = line.employee
        ws.cell(row=row, column=1, value=emp.employee_number or "")
        ws.cell(row=row, column=2, value=emp.name)
        ws.cell(row=row, column=3, value=line.working_days)
        for col_idx, val in enumerate([
            float(line.basic), float(line.allowances), float(line.overtime),
            float(line.bonus), float(line.deductions), float(line.absence_deduction),
            float(line.late_deduction), float(line.advance_deduction), float(line.net),
        ], start=4):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.number_format = "#,##0.00"
        row += 1

    # Totals row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="2563EB")
    for col_idx in range(2, 12):
        ws.cell(row=row, column=col_idx).fill = PatternFill("solid", fgColor="2563EB")
    net_cell = ws.cell(row=row, column=12, value=float(run.total_net))
    net_cell.font = Font(bold=True, color="FFFFFF")
    net_cell.fill = PatternFill("solid", fgColor="2563EB")
    net_cell.number_format = "#,##0.00"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── Generic helpers ────────────────────────────────────────────────────

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _list_pdf(company, title, period_label, headers, rows, totals_row=None, col_widths=None):
    """Generic list-style PDF: title, table, optional totals row.

    headers: list of (label, align) where align is 'left'|'right'
    rows: list of lists matching headers length; values formatted as strings
    totals_row: same shape as a row, displayed bold on a navy background
    col_widths: list of cm widths matching headers length (defaults to even split)
    """
    buf = io.BytesIO()
    p = canvas.Canvas(buf, pagesize=A4)
    _pdf_header(p, company, title, period_label)

    n = len(headers)
    if not col_widths:
        col_widths = [19.0 / n] * n
    # Compute right-edge x for each column (rtl-ish from left edge)
    x_starts = [1 * cm]
    for w in col_widths[:-1]:
        x_starts.append(x_starts[-1] + w * cm)

    y = 24.5 * cm
    # Column header
    p.setFillColor(NAVY)
    p.rect(1 * cm, y - 0.4 * cm, 19 * cm, 0.7 * cm, fill=1, stroke=0)
    p.setFillColor(colors.white)
    p.setFont(_FONT_BOLD, 9)
    for i, (label, align) in enumerate(headers):
        if align == "right":
            p.drawRightString(x_starts[i] + col_widths[i] * cm - 0.2 * cm, y - 0.2 * cm, ar(label))
        else:
            p.drawString(x_starts[i] + 0.2 * cm, y - 0.2 * cm, ar(label))
    y -= 0.9 * cm

    p.setFillColor(colors.black)
    p.setFont(_FONT_REGULAR, 9)
    for row in rows:
        if y < 3 * cm:
            p.showPage()
            _pdf_header(p, company, title + " (cont.)", period_label)
            y = 24.5 * cm
        for i, val in enumerate(row):
            align = headers[i][1]
            if align == "right":
                p.drawRightString(x_starts[i] + col_widths[i] * cm - 0.2 * cm, y, ar(val))
            else:
                p.drawString(x_starts[i] + 0.2 * cm, y, ar(val))
        y -= 0.5 * cm

    if totals_row:
        y -= 0.3 * cm
        p.setFillColor(NAVY)
        p.rect(1 * cm, y - 0.5 * cm, 19 * cm, 0.8 * cm, fill=1, stroke=0)
        p.setFillColor(colors.white)
        p.setFont(_FONT_BOLD, 10)
        for i, val in enumerate(totals_row):
            align = headers[i][1]
            if align == "right":
                p.drawRightString(x_starts[i] + col_widths[i] * cm - 0.2 * cm, y - 0.15 * cm, ar(val))
            else:
                p.drawString(x_starts[i] + 0.2 * cm, y - 0.15 * cm, ar(val))

    p.showPage()
    p.save()
    buf.seek(0)
    return buf


def _list_excel(company, title, period_label, headers, rows, totals_row=None):
    """Generic list-style Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    _excel_styled_header(ws, title, company.name, period_label)

    row_n = 5
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row_n, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0A2540")
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[chr(64 + col)].width = max(14, len(str(h)) + 4)

    row_n += 1
    for row in rows:
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=row_n, column=col, value=val)
            if isinstance(val, (int, float)):
                cell.number_format = "#,##0.00"
        row_n += 1

    if totals_row:
        for col, val in enumerate(totals_row, 1):
            cell = ws.cell(row=row_n, column=col, value=val)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2563EB")
            if isinstance(val, (int, float)):
                cell.number_format = "#,##0.00"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── Income Summary ─────────────────────────────────────────────────────
def export_income_summary(company, fmt, start, end):
    data = income_summary(company.id, start_date=start, end_date=end)
    headers = [("Code", "left"), ("Account", "left"), ("Amount", "right")]
    rows = [[r["code"], r["name"], f"{r['balance']:,.2f}"] for r in data["rows"]]
    totals = ["", "TOTAL REVENUE", f"{data['total']:,.2f}"]
    period = f"{start} → {end}"
    if fmt == "pdf":
        return _list_pdf(company, "Income Summary", period, headers, rows, totals,
                         col_widths=[3, 12, 4]), f"income-summary-{start}-{end}.pdf", "application/pdf"
    return _list_excel(company, "Income Summary", period,
                       ["الكود", "الحساب", "المبلغ"],
                       [[r["code"], r["name"], r["balance"]] for r in data["rows"]],
                       ["", "إجمالي الإيرادات", data["total"]]), \
        f"income-summary-{start}-{end}.xlsx", XLSX_MIME


# ─── Cash Flow ──────────────────────────────────────────────────────────
def export_cash_flow(company, fmt, start, end):
    data = cash_flow(company.id, start_date=start, end_date=end)
    period = f"{start} → {end}" if start else f"→ {end}"
    if fmt == "pdf":
        headers = [("Activity", "left"), ("Net Cash Flow", "right")]
        rows = [
            ["Operating Activities", f"{data['operating']:,.2f}"],
            ["Investing Activities", f"{data['investing']:,.2f}"],
            ["Financing Activities", f"{data['financing']:,.2f}"],
        ]
        totals = ["Net Change in Cash", f"{data['net_change']:,.2f}"]
        return _list_pdf(company, "Cash Flow Statement", period, headers, rows, totals,
                         col_widths=[10, 6]), f"cash-flow-{start}-{end}.pdf", "application/pdf"
    return _list_excel(company, "Cash Flow Statement", period,
                       ["النشاط", "صافي التدفق النقدي"],
                       [["الأنشطة التشغيلية", data["operating"]],
                        ["الأنشطة الاستثمارية", data["investing"]],
                        ["الأنشطة التمويلية", data["financing"]]],
                       ["صافي التغير في النقد", data["net_change"]]), \
        f"cash-flow-{start}-{end}.xlsx", XLSX_MIME


# ─── Expenses Summary ───────────────────────────────────────────────────
def export_expenses_summary(company, fmt, start, end):
    data = expenses_summary(company.id, start_date=start, end_date=end)
    period = f"{start} → {end}"
    if fmt == "pdf":
        headers = [("Code", "left"), ("Account", "left"), ("Entries", "right"), ("Amount", "right")]
        rows = [[r["code"], r["name"], str(r["entry_count"]), f"{r['balance']:,.2f}"] for r in data["rows"]]
        totals = ["", "TOTAL EXPENSES", "", f"{data['total']:,.2f}"]
        return _list_pdf(company, "Expenses Summary", period, headers, rows, totals,
                         col_widths=[3, 10, 2, 4]), f"expenses-summary-{start}-{end}.pdf", "application/pdf"
    return _list_excel(company, "Expenses Summary", period,
                       ["الكود", "الحساب", "عدد القيود", "المبلغ"],
                       [[r["code"], r["name"], r["entry_count"], r["balance"]] for r in data["rows"]],
                       ["", "إجمالي المصروفات", "", data["total"]]), \
        f"expenses-summary-{start}-{end}.xlsx", XLSX_MIME


# ─── P&L Compared ───────────────────────────────────────────────────────
def export_pl_compared(company, fmt, start, end):
    data = income_statement_compared(company.id, start_date=start, end_date=end)
    cur_label = f"{start.isoformat()} → {end.isoformat()}"
    prior_label = f"{data['prior_start'].isoformat()} → {data['prior_end'].isoformat()}"
    period = f"Compared: {cur_label} vs {prior_label}"
    rows_data = [
        ("Revenue", data["current"]["total_revenue"], data["prior"]["total_revenue"], data["delta_revenue"]),
        ("Expenses", data["current"]["total_expense"], data["prior"]["total_expense"], data["delta_expense"]),
        ("Net Profit", data["current"]["net_income"], data["prior"]["net_income"], data["delta_net"]),
    ]
    if fmt == "pdf":
        headers = [("Line", "left"), (cur_label, "right"), (prior_label, "right"), ("Δ", "right")]
        rows = [[label, f"{cur:,.2f}", f"{pri:,.2f}", f"{delta:+,.2f}"] for label, cur, pri, delta in rows_data]
        return _list_pdf(company, "P&L Compared (vs Prior Year)", period, headers, rows,
                         col_widths=[5, 5, 5, 4]), f"pl-compared-{start}-{end}.pdf", "application/pdf"
    return _list_excel(company, "P&L Compared", period,
                       ["البند", "الفترة الحالية", "الفترة السابقة", "التغير"],
                       [[label, cur, pri, delta] for label, cur, pri, delta in rows_data]), \
        f"pl-compared-{start}-{end}.xlsx", XLSX_MIME


# ─── AR Aging ───────────────────────────────────────────────────────────
def export_ar_aging(company, fmt, end):
    data = aging_report(company.id, as_of=end)
    rows_data = [(r["customer"], r["current"], r["d30"], r["d60"], r["d90"], r["d90plus"], r["total"])
                 for r in data["rows"]]
    t = data["totals"]
    period = f"As of {end}"
    if fmt == "pdf":
        headers = [("Customer", "left"), ("Current", "right"), ("1-30", "right"),
                   ("31-60", "right"), ("61-90", "right"), ("90+", "right"), ("Total", "right")]
        rows = [[name, f"{c:,.2f}", f"{d30:,.2f}", f"{d60:,.2f}", f"{d90:,.2f}", f"{d90p:,.2f}", f"{tot:,.2f}"]
                for name, c, d30, d60, d90, d90p, tot in rows_data]
        totals = ["TOTAL", f"{t['current']:,.2f}", f"{t['d30']:,.2f}", f"{t['d60']:,.2f}",
                  f"{t['d90']:,.2f}", f"{t['d90plus']:,.2f}", f"{t['total']:,.2f}"]
        return _list_pdf(company, "Accounts Receivable Aging", period, headers, rows, totals,
                         col_widths=[5, 2.5, 2.5, 2.5, 2.5, 2.5, 1.5]), \
            f"ar-aging-{end}.pdf", "application/pdf"
    return _list_excel(company, "AR Aging", period,
                       ["العميل", "جاري", "1-30", "31-60", "61-90", "90+", "الإجمالي"],
                       [list(r) for r in rows_data],
                       ["الإجمالي", t["current"], t["d30"], t["d60"], t["d90"], t["d90plus"], t["total"]]), \
        f"ar-aging-{end}.xlsx", XLSX_MIME


# ─── AP Aging ───────────────────────────────────────────────────────────
def export_ap_aging(company, fmt, end):
    data = ap_aging_report(company.id, as_of=end)
    rows_data = [(r["vendor"], r["current"], r["d30"], r["d60"], r["d90"], r["d90plus"], r["total"])
                 for r in data["rows"]]
    t = data["totals"]
    period = f"As of {end}"
    if fmt == "pdf":
        headers = [("Vendor", "left"), ("Current", "right"), ("1-30", "right"),
                   ("31-60", "right"), ("61-90", "right"), ("90+", "right"), ("Total", "right")]
        rows = [[name, f"{c:,.2f}", f"{d30:,.2f}", f"{d60:,.2f}", f"{d90:,.2f}", f"{d90p:,.2f}", f"{tot:,.2f}"]
                for name, c, d30, d60, d90, d90p, tot in rows_data]
        totals = ["TOTAL", f"{t['current']:,.2f}", f"{t['d30']:,.2f}", f"{t['d60']:,.2f}",
                  f"{t['d90']:,.2f}", f"{t['d90plus']:,.2f}", f"{t['total']:,.2f}"]
        return _list_pdf(company, "Accounts Payable Aging", period, headers, rows, totals,
                         col_widths=[5, 2.5, 2.5, 2.5, 2.5, 2.5, 1.5]), \
            f"ap-aging-{end}.pdf", "application/pdf"
    return _list_excel(company, "AP Aging", period,
                       ["المورد", "جاري", "1-30", "31-60", "61-90", "90+", "الإجمالي"],
                       [list(r) for r in rows_data],
                       ["الإجمالي", t["current"], t["d30"], t["d60"], t["d90"], t["d90plus"], t["total"]]), \
        f"ap-aging-{end}.xlsx", XLSX_MIME


# ─── VAT Report ─────────────────────────────────────────────────────────
def export_vat_report(company, fmt, start, end):
    data = vat_report(company.id, start_date=start, end_date=end)
    period = f"{start} → {end}"
    if fmt == "pdf":
        # Special gov-ready layout — single-page summary
        buf = io.BytesIO()
        p = canvas.Canvas(buf, pagesize=A4)
        _pdf_header(p, company, "VAT Return", period)
        y = 24 * cm
        p.setFillColor(colors.HexColor("#475569"))
        p.setFont(_FONT_REGULAR, 11)
        if company.tax_number:
            p.drawString(1.5 * cm, y, ar(f"Tax Number: {company.tax_number}"))
            y -= 0.8 * cm

        # Three rows: collected, paid, net — large summary cards
        for label, amount, color in [
            ("VAT Collected (from sales)", data["collected"], colors.HexColor("#10B981")),
            ("VAT Paid (to suppliers)", data["paid"], colors.HexColor("#F59E0B")),
        ]:
            p.setFillColor(color)
            p.rect(1 * cm, y - 1.5 * cm, 19 * cm, 1.2 * cm, fill=1, stroke=0)
            p.setFillColor(colors.white)
            p.setFont(_FONT_BOLD, 12)
            p.drawString(1.5 * cm, y - 0.8 * cm, ar(label))
            p.setFont(_FONT_BOLD, 16)
            p.drawRightString(19.5 * cm, y - 0.8 * cm, f"{amount:,.2f} {company.base_currency}")
            y -= 2 * cm

        p.setFillColor(NAVY)
        p.rect(1 * cm, y - 1.5 * cm, 19 * cm, 1.4 * cm, fill=1, stroke=0)
        p.setFillColor(colors.white)
        p.setFont(_FONT_BOLD, 13)
        p.drawString(1.5 * cm, y - 0.8 * cm, ar("NET DUE TO GOVERNMENT"))
        p.setFont(_FONT_BOLD, 18)
        p.drawRightString(19.5 * cm, y - 0.8 * cm, f"{data['net']:,.2f} {company.base_currency}")

        p.showPage()
        p.save()
        buf.seek(0)
        return buf, f"vat-{start}-{end}.pdf", "application/pdf"

    return _list_excel(company, "VAT Report", period,
                       ["البند", "المبلغ"],
                       [["الضريبة المحصلة", data["collected"]],
                        ["الضريبة المدفوعة", data["paid"]],
                        ["الصافي المستحق", data["net"]]]), \
        f"vat-{start}-{end}.xlsx", XLSX_MIME


# ─── Payroll Summary ────────────────────────────────────────────────────
def export_payroll_summary(company, fmt, year=None, month=None):
    data = payroll_summary_report(company.id, year=year, month=month)
    rows_data = [(r["period"], r["run_number"], r["employee"], r["basic"], r["allowances"],
                  r["overtime"], r["bonus"], r["deductions"], r["net"]) for r in data["rows"]]
    t = data["totals"]
    period = f"Year {year or 'all'} · Month {month or 'all'}"
    if fmt == "pdf":
        headers = [("Period", "left"), ("Run", "left"), ("Employee", "left"),
                   ("Basic", "right"), ("Allow", "right"), ("OT", "right"),
                   ("Bonus", "right"), ("Deduct", "right"), ("Net", "right")]
        rows = [[per, run, emp, f"{b:,.2f}", f"{a:,.2f}", f"{ot:,.2f}", f"{bn:,.2f}", f"{d:,.2f}", f"{n:,.2f}"]
                for per, run, emp, b, a, ot, bn, d, n in rows_data]
        totals = ["", "", "TOTAL", f"{t['basic']:,.2f}", f"{t['allowances']:,.2f}",
                  f"{t['overtime']:,.2f}", f"{t['bonus']:,.2f}", f"{t['deductions']:,.2f}", f"{t['net']:,.2f}"]
        return _list_pdf(company, "Payroll Summary", period, headers, rows, totals,
                         col_widths=[1.5, 2, 3.5, 2, 2, 1.5, 1.5, 2, 2]), \
            f"payroll-summary.pdf", "application/pdf"
    return _list_excel(company, "Payroll Summary", period,
                       ["الفترة", "الكشف", "الموظف", "الأساسي", "البدلات", "أوفرتايم", "بونص", "خصومات", "الصافي"],
                       [list(r) for r in rows_data],
                       ["", "", "الإجمالي", t["basic"], t["allowances"], t["overtime"], t["bonus"], t["deductions"], t["net"]]), \
        f"payroll-summary.xlsx", XLSX_MIME


# ─── Fixed Assets Report ────────────────────────────────────────────────
def export_fixed_assets(company, fmt):
    data = fixed_assets_report(company.id)
    rows_data = [(r["name"], r["vendor"] or "—", str(r["purchase_date"]),
                  f"{r['useful_life_years']}y", r["cost"], r["annual_dep"],
                  r["accumulated_dep"], r["nbv"]) for r in data["rows"]]
    t = data["totals"]
    period = "All active fixed assets"
    if fmt == "pdf":
        headers = [("Asset", "left"), ("Vendor", "left"), ("Purchase", "left"),
                   ("Life", "left"), ("Cost", "right"), ("Annual Dep", "right"),
                   ("Acc. Dep", "right"), ("NBV", "right")]
        rows = [[n, v, p, l, f"{c:,.2f}", f"{ad:,.2f}", f"{accd:,.2f}", f"{nbv:,.2f}"]
                for n, v, p, l, c, ad, accd, nbv in rows_data]
        totals = ["TOTAL", "", "", "", f"{t['cost']:,.2f}", f"{t['annual_dep']:,.2f}",
                  f"{t['accumulated_dep']:,.2f}", f"{t['nbv']:,.2f}"]
        return _list_pdf(company, "Fixed Assets Report", period, headers, rows, totals,
                         col_widths=[3, 2.5, 2, 1.2, 2.5, 2.5, 2.5, 2.8]), \
            f"fixed-assets.pdf", "application/pdf"
    return _list_excel(company, "Fixed Assets", period,
                       ["الأصل", "المورد", "تاريخ الشراء", "العمر", "التكلفة", "إهلاك سنوي", "مجمع الإهلاك", "القيمة الدفترية"],
                       [list(r) for r in rows_data],
                       ["الإجمالي", "", "", "", t["cost"], t["annual_dep"], t["accumulated_dep"], t["nbv"]]), \
        f"fixed-assets.xlsx", XLSX_MIME


# ─── Dispatcher ─────────────────────────────────────────────────────────
def export_report(company, report_type, fmt, start, end, **kwargs):
    """Dispatch to the right export function. fmt = 'pdf' or 'excel'."""
    if report_type == "balance-sheet":
        if fmt == "pdf":
            return export_balance_sheet_pdf(company, end), f"balance-sheet-{end}.pdf", "application/pdf"
        return export_balance_sheet_excel(company, end), f"balance-sheet-{end}.xlsx", XLSX_MIME
    if report_type == "income-statement":
        if fmt == "pdf":
            return export_income_statement_pdf(company, start, end), f"income-statement-{start}-{end}.pdf", "application/pdf"
        return export_income_statement_excel(company, start, end), f"income-statement-{start}-{end}.xlsx", XLSX_MIME
    if report_type == "cash-flow":
        return export_cash_flow(company, fmt, start, end)
    if report_type == "income-summary":
        return export_income_summary(company, fmt, start, end)
    if report_type == "expenses-summary":
        return export_expenses_summary(company, fmt, start, end)
    if report_type == "pl-compared":
        return export_pl_compared(company, fmt, start, end)
    if report_type == "ar-aging":
        return export_ar_aging(company, fmt, end)
    if report_type == "ap-aging":
        return export_ap_aging(company, fmt, end)
    if report_type == "vat":
        return export_vat_report(company, fmt, start, end)
    if report_type == "payroll-summary":
        return export_payroll_summary(company, fmt,
                                      year=kwargs.get("year"), month=kwargs.get("month"))
    if report_type == "fixed-assets":
        return export_fixed_assets(company, fmt)
    raise ValueError(f"Unknown report: {report_type}")
