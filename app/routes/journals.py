from datetime import datetime, date
from flask import Blueprint, render_template, redirect, url_for, flash, request, g, send_file
from flask_login import login_required, current_user
from sqlalchemy import func, or_
from app import db
from app.models import (
    JournalEntry, JournalLine, Account, Customer, Vendor,
    JournalAudit, JournalAction, JournalTemplate, JournalTemplateLine,
    RecurringJournal, RecurrenceFrequency,
)
from app.services.ledger import post_journal, reverse_journal, LedgerError
from app.services.journals import (
    pause_entry, reactivate_entry, post_from_template,
)

bp = Blueprint("journals", __name__)


def _parse_date(s, default=None):
    if not s:
        return default
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return default


@bp.route("/")
@login_required
def index():
    if not g.active_company:
        return redirect(url_for("companies.new"))

    q = JournalEntry.query.filter_by(company_id=g.active_company.id)

    # Search: number / description / reference / amount / customer / vendor
    search = (request.args.get("search") or "").strip()
    if search:
        like = f"%{search}%"
        amount_search = None
        try:
            amount_search = float(search)
        except ValueError:
            pass

        # Customer + vendor names: match invoice/payment journals where the related
        # invoice's customer name matches, or vendor-sourced journals
        from app.models import Invoice
        customer_match = db.session.query(Invoice.id).join(
            Customer, Invoice.customer_id == Customer.id
        ).filter(Customer.name.ilike(like)).subquery()
        vendor_match = db.session.query(Vendor.id).filter(Vendor.name.ilike(like)).subquery()

        clauses = [
            JournalEntry.number.ilike(like),
            JournalEntry.description.ilike(like),
            JournalEntry.reference.ilike(like),
            db.and_(
                JournalEntry.source_type.in_(["invoice", "payment", "refund", "credit_note"]),
                JournalEntry.source_id.in_(customer_match),
            ),
            db.and_(
                JournalEntry.source_type == "vendor_payment",
                JournalEntry.source_id.in_(vendor_match),
            ),
        ]
        if amount_search is not None:
            matching_ids = db.session.query(JournalLine.entry_id).filter(
                or_(JournalLine.debit == amount_search, JournalLine.credit == amount_search)
            ).distinct().subquery()
            clauses.append(JournalEntry.id.in_(matching_ids))
        q = q.filter(or_(*clauses))

    # Amount range filter (from/to)
    amount_min = request.args.get("amount_min")
    amount_max = request.args.get("amount_max")
    if amount_min or amount_max:
        line_q = db.session.query(JournalLine.entry_id).distinct()
        if amount_min:
            try:
                amn = float(amount_min)
                line_q = line_q.filter(or_(JournalLine.debit >= amn, JournalLine.credit >= amn))
            except ValueError:
                pass
        if amount_max:
            try:
                amx = float(amount_max)
                line_q = line_q.filter(or_(
                    db.and_(JournalLine.debit > 0, JournalLine.debit <= amx),
                    db.and_(JournalLine.credit > 0, JournalLine.credit <= amx),
                ))
            except ValueError:
                pass
        q = q.filter(JournalEntry.id.in_(line_q.subquery()))

    # Filters
    start_date = _parse_date(request.args.get("start_date"))
    end_date = _parse_date(request.args.get("end_date"))
    if start_date:
        q = q.filter(JournalEntry.date >= start_date)
    if end_date:
        q = q.filter(JournalEntry.date <= end_date)

    source_type = request.args.get("source_type")
    if source_type:
        if source_type == "manual":
            q = q.filter(JournalEntry.source_type.is_(None))
        elif source_type == "reversal":
            q = q.filter(JournalEntry.is_reversal.is_(True))
        else:
            q = q.filter(JournalEntry.source_type == source_type)

    status = request.args.get("status")
    if status == "active":
        q = q.filter(JournalEntry.is_active.is_(True))
    elif status == "paused":
        q = q.filter(JournalEntry.is_active.is_(False))

    account_id = request.args.get("account_id")
    if account_id:
        matching = db.session.query(JournalLine.entry_id).filter(
            JournalLine.account_id == int(account_id)
        ).distinct().subquery()
        q = q.filter(JournalEntry.id.in_(matching))

    user_id = request.args.get("user_id")
    if user_id:
        q = q.filter(JournalEntry.created_by == int(user_id))

    reference_prefix = (request.args.get("reference") or "").strip()
    if reference_prefix:
        q = q.filter(JournalEntry.reference.ilike(f"{reference_prefix}%"))

    # Sort
    sort = request.args.get("sort", "date_desc")
    sort_map = {
        "date_desc": (JournalEntry.date.desc(), JournalEntry.id.desc()),
        "date_asc": (JournalEntry.date.asc(), JournalEntry.id.asc()),
        "number_desc": (JournalEntry.id.desc(),),
        "number_asc": (JournalEntry.id.asc(),),
    }
    for clause in sort_map.get(sort, sort_map["date_desc"]):
        q = q.order_by(clause)

    # Pagination
    page = max(int(request.args.get("page", 1)), 1)
    per_page = 25
    total_count = q.count()
    entries = q.offset((page - 1) * per_page).limit(per_page).all()
    total_pages = max((total_count + per_page - 1) // per_page, 1)

    # Filtered totals (across ALL filtered rows, not just this page)
    total_debit_q = db.session.query(func.coalesce(func.sum(JournalLine.debit), 0)).filter(
        JournalLine.entry_id.in_(q.with_entities(JournalEntry.id).subquery())
    ).scalar()
    total_credit_q = db.session.query(func.coalesce(func.sum(JournalLine.credit), 0)).filter(
        JournalLine.entry_id.in_(q.with_entities(JournalEntry.id).subquery())
    ).scalar()

    accounts = Account.query.filter_by(
        company_id=g.active_company.id, is_active=True
    ).order_by(Account.code).all()

    # Users who have ever created an entry in this company — for the user filter dropdown
    from app.models import User as UserModel
    user_ids = db.session.query(JournalEntry.created_by).filter(
        JournalEntry.company_id == g.active_company.id,
        JournalEntry.created_by.isnot(None),
    ).distinct().all()
    users = UserModel.query.filter(UserModel.id.in_([u[0] for u in user_ids])).all()

    # If this is an HTMX partial request, return only the results fragment
    if request.headers.get("HX-Request"):
        return render_template(
            "journals/_results.html",
            entries=entries, total_count=total_count, page=page, total_pages=total_pages,
            total_debit=float(total_debit_q or 0), total_credit=float(total_credit_q or 0),
            sort=sort,
        )

    return render_template(
        "journals/index.html",
        entries=entries,
        total_count=total_count, page=page, total_pages=total_pages,
        total_debit=float(total_debit_q or 0),
        total_credit=float(total_credit_q or 0),
        accounts=accounts, users=users,
        sort=sort,
    )


@bp.route("/new", methods=["GET", "POST"])
@login_required
def new():
    if not g.active_company:
        return redirect(url_for("companies.new"))
    accounts = Account.query.filter_by(company_id=g.active_company.id, is_active=True).order_by(Account.code).all()
    templates = JournalTemplate.query.filter_by(
        company_id=g.active_company.id, is_active=True
    ).order_by(JournalTemplate.name).all()

    if request.method == "POST":
        description = request.form.get("description", "").strip()
        reference = request.form.get("reference", "").strip() or None
        entry_date = _parse_date(request.form.get("date"), date.today())
        currency = request.form.get("currency", g.active_company.base_currency)
        exchange_rate = float(request.form.get("exchange_rate", 1.0))

        account_ids = request.form.getlist("account_id[]")
        debits = request.form.getlist("debit[]")
        credits = request.form.getlist("credit[]")
        memos = request.form.getlist("memo[]")

        lines = []
        for i, aid in enumerate(account_ids):
            if not aid:
                continue
            d = float(debits[i] or 0)
            c = float(credits[i] or 0)
            if d == 0 and c == 0:
                continue
            lines.append({
                "account_id": int(aid), "debit": d, "credit": c,
                "memo": memos[i] if i < len(memos) else None,
            })

        try:
            entry = post_journal(
                company_id=g.active_company.id,
                description=description, lines=lines,
                entry_date=entry_date, reference=reference,
                currency=currency, exchange_rate=exchange_rate,
                created_by=current_user.id,
            )
            db.session.add(JournalAudit(
                entry_id=entry.id, user_id=current_user.id,
                action=JournalAction.CREATED, reason="إنشاء يدوي",
            ))
            db.session.commit()
            flash(f"تم تسجيل القيد {entry.number}", "success")
            return redirect(url_for("journals.view", entry_id=entry.id))
        except LedgerError as e:
            flash(str(e), "error")

    return render_template("journals/form.html", accounts=accounts, templates=templates)


@bp.route("/<int:entry_id>")
@login_required
def view(entry_id):
    entry = db.session.get(JournalEntry, entry_id)
    if not entry or entry.company_id != g.active_company.id:
        flash("غير موجود", "error")
        return redirect(url_for("journals.index"))

    # Resolve source link if available
    source_link = None
    source_label = None
    if entry.source_type == "invoice" and entry.source_id:
        from app.models import Invoice
        inv = db.session.get(Invoice, entry.source_id)
        if inv and inv.company_id == g.active_company.id:
            source_link = url_for("invoices.view", invoice_id=inv.id)
            source_label = f"فاتورة {inv.number}"
    elif entry.source_type == "payment" and entry.source_id:
        from app.models import Invoice
        inv = db.session.get(Invoice, entry.source_id)
        if inv and inv.company_id == g.active_company.id:
            source_link = url_for("invoices.view", invoice_id=inv.id)
            source_label = f"دفعة على فاتورة {inv.number}"
    elif entry.source_type == "payroll" and entry.source_id:
        from app.models import PayrollRun
        run = db.session.get(PayrollRun, entry.source_id)
        if run and run.company_id == g.active_company.id:
            source_link = url_for("payroll.view_run", run_id=run.id)
            source_label = f"كشف رواتب {run.number or ''}"

    audits = JournalAudit.query.filter_by(entry_id=entry.id).order_by(JournalAudit.created_at).all()
    return render_template(
        "journals/view.html",
        entry=entry, source_link=source_link, source_label=source_label, audits=audits,
    )


@bp.route("/<int:entry_id>/reverse", methods=["POST"])
@login_required
def reverse(entry_id):
    entry = db.session.get(JournalEntry, entry_id)
    if not entry or entry.company_id != g.active_company.id:
        flash("غير موجود", "error")
        return redirect(url_for("journals.index"))
    try:
        new_entry = reverse_journal(entry_id, created_by=current_user.id)
        db.session.add(JournalAudit(
            entry_id=entry.id, user_id=current_user.id,
            action=JournalAction.REVERSED, reason=f"تم إنشاء قيد عكسي {new_entry.number}",
        ))
        db.session.commit()
        flash(f"تم إنشاء قيد عكسي {new_entry.number}", "success")
    except LedgerError as e:
        flash(str(e), "error")
    return redirect(url_for("journals.view", entry_id=entry_id))


@bp.route("/<int:entry_id>/pause", methods=["POST"])
@login_required
def pause(entry_id):
    entry = db.session.get(JournalEntry, entry_id)
    if not entry or entry.company_id != g.active_company.id:
        flash("غير موجود", "error")
        return redirect(url_for("journals.index"))
    try:
        pause_entry(entry, request.form.get("reason", ""), current_user.id)
        flash("تم إيقاف القيد — لا يؤثر على التقارير", "success")
    except LedgerError as e:
        flash(str(e), "error")
    return redirect(url_for("journals.view", entry_id=entry_id))


@bp.route("/<int:entry_id>/reactivate", methods=["POST"])
@login_required
def reactivate(entry_id):
    entry = db.session.get(JournalEntry, entry_id)
    if not entry or entry.company_id != g.active_company.id:
        flash("غير موجود", "error")
        return redirect(url_for("journals.index"))
    try:
        reactivate_entry(entry, request.form.get("reason", ""), current_user.id)
        flash("تم إعادة تنشيط القيد — التقارير محدّثة", "success")
    except LedgerError as e:
        flash(str(e), "error")
    return redirect(url_for("journals.view", entry_id=entry_id))


@bp.route("/templates")
@login_required
def templates_list():
    templates = JournalTemplate.query.filter_by(company_id=g.active_company.id).all()
    return render_template("journals/templates.html", templates=templates)


@bp.route("/templates/new", methods=["GET", "POST"])
@login_required
def templates_new():
    accounts = Account.query.filter_by(
        company_id=g.active_company.id, is_active=True
    ).order_by(Account.code).all()
    if request.method == "POST":
        try:
            tpl = JournalTemplate(
                company_id=g.active_company.id,
                name=request.form.get("name", "").strip(),
                description=request.form.get("description", ""),
            )
            if not tpl.name:
                raise ValueError("اسم القالب مطلوب")
            db.session.add(tpl)
            db.session.flush()
            account_ids = request.form.getlist("account_id[]")
            debits = request.form.getlist("debit[]")
            credits = request.form.getlist("credit[]")
            memos = request.form.getlist("memo[]")
            for i, aid in enumerate(account_ids):
                if not aid:
                    continue
                d, c = float(debits[i] or 0), float(credits[i] or 0)
                if d == 0 and c == 0:
                    continue
                db.session.add(JournalTemplateLine(
                    template_id=tpl.id, account_id=int(aid),
                    debit=d, credit=c,
                    memo=memos[i] if i < len(memos) else None,
                ))
            db.session.commit()
            flash("تم حفظ القالب", "success")
            return redirect(url_for("journals.templates_list"))
        except ValueError as e:
            flash(str(e), "error")
    return render_template("journals/template_form.html", accounts=accounts)


@bp.route("/templates/<int:template_id>/use")
@login_required
def use_template(template_id):
    tpl = db.session.get(JournalTemplate, template_id)
    if not tpl or tpl.company_id != g.active_company.id:
        return redirect(url_for("journals.templates_list"))
    accounts = Account.query.filter_by(
        company_id=g.active_company.id, is_active=True
    ).order_by(Account.code).all()
    templates = JournalTemplate.query.filter_by(company_id=g.active_company.id, is_active=True).all()
    return render_template("journals/form.html", accounts=accounts, templates=templates, prefill_template=tpl)


@bp.route("/recurring")
@login_required
def recurring_list():
    items = RecurringJournal.query.filter_by(company_id=g.active_company.id).all()
    return render_template("journals/recurring.html", items=items)


@bp.route("/recurring/new", methods=["GET", "POST"])
@login_required
def recurring_new():
    templates = JournalTemplate.query.filter_by(
        company_id=g.active_company.id, is_active=True
    ).all()
    if request.method == "POST":
        try:
            r = RecurringJournal(
                company_id=g.active_company.id,
                template_id=int(request.form.get("template_id")),
                name=request.form.get("name", "").strip(),
                frequency=RecurrenceFrequency[request.form.get("frequency")],
                next_run_date=_parse_date(request.form.get("next_run_date"), date.today()),
                end_date=_parse_date(request.form.get("end_date")),
            )
            if not r.name:
                raise ValueError("الاسم مطلوب")
            db.session.add(r)
            db.session.commit()
            flash("تم إنشاء الجدول التلقائي", "success")
            return redirect(url_for("journals.recurring_list"))
        except (ValueError, KeyError) as e:
            flash(str(e), "error")
    return render_template("journals/recurring_form.html", templates=templates)


def _build_filtered_query(company_id, args):
    """Shared query builder so per-entry and bulk endpoints respect the same filters."""
    q = JournalEntry.query.filter_by(company_id=company_id)
    sd = _parse_date(args.get("start_date"))
    ed = _parse_date(args.get("end_date"))
    if sd: q = q.filter(JournalEntry.date >= sd)
    if ed: q = q.filter(JournalEntry.date <= ed)
    source_type = args.get("source_type")
    if source_type == "manual":
        q = q.filter(JournalEntry.source_type.is_(None))
    elif source_type == "reversal":
        q = q.filter(JournalEntry.is_reversal.is_(True))
    elif source_type:
        q = q.filter(JournalEntry.source_type == source_type)
    status = args.get("status")
    if status == "active":
        q = q.filter(JournalEntry.is_active.is_(True))
    elif status == "paused":
        q = q.filter(JournalEntry.is_active.is_(False))
    return q


@bp.route("/export")
@login_required
def export_filtered():
    """Export the currently filtered list as Excel or PDF (fmt=pdf|excel)."""
    fmt = request.args.get("fmt", "excel")
    q = _build_filtered_query(g.active_company.id, request.args)
    entries = q.order_by(JournalEntry.date.desc()).all()

    if fmt == "pdf":
        from app.services.export import export_journals_list_pdf
        period = ""
        if request.args.get("start_date") or request.args.get("end_date"):
            period = f"{request.args.get('start_date', '')} → {request.args.get('end_date', '')}"
        buf = export_journals_list_pdf(g.active_company, entries, period_label=period)
        return send_file(buf, mimetype="application/pdf",
                         as_attachment=True, download_name=f"journals-{date.today()}.pdf")

    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Journals"
    headers = ["الرقم", "التاريخ", "الوصف", "المرجع", "المدين", "الدائن", "الحالة"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0A2540")
    row = 2
    for e in entries:
        ws.cell(row=row, column=1, value=e.number)
        ws.cell(row=row, column=2, value=str(e.date))
        ws.cell(row=row, column=3, value=e.description)
        ws.cell(row=row, column=4, value=e.reference or "")
        ws.cell(row=row, column=5, value=float(e.total_debit)).number_format = "#,##0.00"
        ws.cell(row=row, column=6, value=float(e.total_credit)).number_format = "#,##0.00"
        ws.cell(row=row, column=7, value="نشط" if e.is_active else "موقوف")
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"journals-{date.today()}.xlsx")


@bp.route("/<int:entry_id>/export/<fmt>")
@login_required
def export_entry(entry_id, fmt):
    entry = db.session.get(JournalEntry, entry_id)
    if not entry or entry.company_id != g.active_company.id:
        return redirect(url_for("journals.index"))
    if fmt == "pdf":
        from app.services.export import export_journal_entry_pdf
        buf = export_journal_entry_pdf(entry)
        return send_file(buf, mimetype="application/pdf",
                         download_name=f"{entry.number or 'entry-' + str(entry.id)}.pdf",
                         as_attachment=True)
    if fmt == "excel":
        from app.services.export import export_journal_entry_excel
        buf = export_journal_entry_excel(entry)
        return send_file(buf,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         download_name=f"{entry.number or 'entry-' + str(entry.id)}.xlsx",
                         as_attachment=True)
    return redirect(url_for("journals.view", entry_id=entry.id))


@bp.route("/bulk", methods=["POST"])
@login_required
def bulk_action():
    """Handle bulk actions on selected entries: export, pause."""
    action = request.form.get("action")
    selected_ids = request.form.getlist("entry_ids[]")
    if not selected_ids:
        flash("لم يتم تحديد أي قيد", "warning")
        return redirect(url_for("journals.index"))

    entries = JournalEntry.query.filter(
        JournalEntry.id.in_([int(i) for i in selected_ids]),
        JournalEntry.company_id == g.active_company.id,
    ).all()

    if action == "export_excel":
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = "Selected Journals"
        for col, h in enumerate(["الرقم", "التاريخ", "الوصف", "المرجع", "المدين", "الدائن", "الحالة"], 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="0A2540")
        row = 2
        for e in entries:
            ws.cell(row=row, column=1, value=e.number)
            ws.cell(row=row, column=2, value=str(e.date))
            ws.cell(row=row, column=3, value=e.description)
            ws.cell(row=row, column=4, value=e.reference or "")
            ws.cell(row=row, column=5, value=float(e.total_debit)).number_format = "#,##0.00"
            ws.cell(row=row, column=6, value=float(e.total_credit)).number_format = "#,##0.00"
            ws.cell(row=row, column=7, value="نشط" if e.is_active else "موقوف")
            row += 1
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=f"selected-journals-{date.today()}.xlsx")

    if action == "export_pdf":
        from app.services.export import export_journals_list_pdf
        buf = export_journals_list_pdf(g.active_company, entries,
                                       period_label=f"{len(entries)} selected entries")
        return send_file(buf, mimetype="application/pdf",
                         as_attachment=True, download_name=f"selected-journals-{date.today()}.pdf")

    if action == "pause":
        reason = (request.form.get("bulk_reason") or "").strip()
        if not reason:
            flash("سبب الإيقاف مطلوب للإيقاف الجماعي", "error")
            return redirect(url_for("journals.index"))
        paused_count = 0
        for e in entries:
            try:
                pause_entry(e, reason, current_user.id)
                paused_count += 1
            except LedgerError:
                pass
        flash(f"تم إيقاف {paused_count} قيد من أصل {len(entries)}", "success")
        return redirect(url_for("journals.index"))

    flash("إجراء غير معروف", "error")
    return redirect(url_for("journals.index"))
